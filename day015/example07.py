"""
example07 - 

Author: kaiby
Date: 2024/1/15 17:38
"""
import openpyxl
from openpyxl.styles import Font, Alignment

wb = openpyxl.load_workbook('score.xlsx')
sheet = wb.worksheets[0]
sheet['E1'] = '平均分'
for row_idx in range(2, 7):
    sheet[f'E{row_idx}'] = f'=average(B{row_idx}:D{row_idx})'
    sheet.cell(row_idx, 5).font = Font(size=18, bold=True, color='ff1493', name='微软雅黑')
    sheet.cell(row_idx, 5).alignment = Alignment(horizontal='center', vertical='center')

wb.save('score.xlsx')
