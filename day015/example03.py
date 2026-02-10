"""
example03 - 读取xlsx

Author: kaiby
Date: 2024/1/12 15:59
"""
import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook('id.xlsx')
print(wb.sheetnames)
sheet = wb['tem4']
# sheet = wb.worksheets[0]
print(sheet.dimensions)
print(sheet.max_row, sheet.max_column)

# 循环单元格
# for row in range(1, sheet.max_row + 1):
#     for col in range(1, sheet.max_column + 1):
#         cell = sheet.cell(row, col)
#         if col == 9 and type(cell.value) is datetime:
#             print(cell.value.strftime('%Y年%m月%d日'), end=' ')
#         else:
#             print(cell.value, end=' ')
#     print()

# 另一种方式循环
#for row in sheet.rows:
# 指定从第几行开始
for row in sheet.iter_rows(2):
    for cell in row:
        print(cell.value, end=' ')
    print()

# 另一种方式
print(sheet.cell(3, 2).value)
print(sheet['B3'].value)
