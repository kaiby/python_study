"""
example04 - 

Author: kaiby
Date: 2024/1/15 14:04
"""
import random

import openpyxl

# 创建工作簿
wb = openpyxl.Workbook()

# 创建工作表
# sheet = wb.create_sheet('成绩')
# 拿默认工作表
sheet = wb.active
sheet.title = '成绩'

# 向单元格写入数据
sheet.cell(1, 1, '姓名')
sheet.cell(1, 2, '语文')
sheet.cell(1, 3, '数学')
sheet.cell(1, 4, '英语')

students = ('Tom', 'Jerry', 'Anna', 'Emma', 'Junior')

for i in range(len(students)):
    sheet.cell(i + 2, 1, students[i])
    for col in range(2, 5):
        sheet.cell(i + 2, col, random.randrange(40, 101))


# 保存
wb.save('score.xlsx')
